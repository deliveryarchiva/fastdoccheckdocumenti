"""
Data loader for Archiva File DB.

Sorgenti dati:
  - Postel  → Excel "Estrazione Postel"
               col 0: r_object_id | col 2: object_name | col 32: pt_ragione_sociale | col 88: pt_piva

  - Archiva → due alternative (priorità: CSV > foglio Excel):
      a) CSV fastweb_doc_in_requiro:  col 0: r_object_id | col 9: file_name | col 96: PIVA | col 99: RAGIONE_SOCIALE
      b) Excel "Estrazione Archiva":  col 7: file_name | col 94: PIVA | col 97: RAGIONE_SOCIALE | col 115: r_object_id

Filename date pattern: PIVA_YYYY_M_D_...
"""

import re
import logging
import os
import pickle

logger = logging.getLogger(__name__)
_FLOAT_RE = re.compile(r"^(\d+)\.0$")


def _clean_str(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    m = _FLOAT_RE.match(s)
    if m:
        return m.group(1)
    if s.lower() in ("nan", "none", "nat", "nulldate"):
        return ""
    return s


def parse_date_from_filename(filename: str):
    if not filename:
        return None
    try:
        parts = filename.split("_")
        if len(parts) >= 4:
            year  = int(parts[1])
            month = int(parts[2])
            day   = int(parts[3])
            if 2000 <= year <= 2099 and 1 <= month <= 12 and 1 <= day <= 31:
                from datetime import date
                return date(year, month, day)
    except (ValueError, IndexError):
        pass
    return None


def _mtime(path) -> float:
    try:
        return os.path.getmtime(path) if path and os.path.exists(path) else 0.0
    except OSError:
        return 0.0


# ── Public entry point ────────────────────────────────────────────────────────

def load_data(excel_path: str, csv_path: str = None) -> list:
    import time
    cache_path = excel_path + ".cache.pkl"

    if os.path.exists(cache_path):
        cache_mt = _mtime(cache_path)
        if cache_mt >= _mtime(excel_path) and cache_mt >= _mtime(csv_path):
            try:
                t0 = time.time()
                logger.info(f"Loading from cache …")
                with open(cache_path, "rb") as f:
                    records = pickle.load(f)
                for r in records:
                    if "anno_mese" not in r:
                        dd = r.get("data_doc", "")
                        r["anno_mese"] = dd[:7].replace("-", "/") if len(dd) >= 7 else ""
                logger.info(f"Cache loaded in {time.time()-t0:.1f}s — {len(records):,} records")
                return records
            except Exception as e:
                logger.warning(f"Cache load failed ({e}), reloading")

    rows_postel  = _load_postel(excel_path)
    rows_archiva = _load_archiva_csv(csv_path) \
                   if (csv_path and os.path.exists(csv_path)) \
                   else _load_archiva_excel(excel_path)

    records = _build_records(rows_archiva, rows_postel)

    try:
        with open(cache_path, "wb") as f:
            pickle.dump(records, f, protocol=pickle.HIGHEST_PROTOCOL)
        logger.info(f"Cache saved")
    except Exception as e:
        logger.warning(f"Cache save failed: {e}")

    return records


# ── Loaders ───────────────────────────────────────────────────────────────────

def _load_postel(excel_path: str) -> list:
    """→ list of [r_object_id, nome_file, ragione_sociale, piva]"""
    logger.info("Loading Postel from Excel …")
    try:
        import pandas as pd
        df = pd.read_excel(excel_path, sheet_name="Estrazione Postel",
                           usecols=[0, 2, 32, 88], header=0, dtype=object, engine="openpyxl")
        df.columns = ["r_object_id", "nome_file", "ragione_sociale", "piva"]
        logger.info(f"  Postel: {len(df):,} righe")
        return df.values.tolist()
    except Exception as e:
        logger.warning(f"Postel pandas failed ({e}), openpyxl fallback")
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb["Estrazione Postel"]
        cols = [0, 2, 32, 88]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0: continue
            rows.append([row[c] if c < len(row) else None for c in cols])
        wb.close()
        return rows


def _load_archiva_excel(excel_path: str) -> list:
    """→ list of [nome_file, piva, ragione_sociale, r_object_id]"""
    logger.info("Loading Archiva from Excel sheet …")
    try:
        import pandas as pd
        df = pd.read_excel(excel_path, sheet_name="Estrazione Archiva",
                           usecols=[7, 94, 97, 115], header=0, dtype=object, engine="openpyxl")
        df.columns = ["nome_file", "piva", "ragione_sociale", "r_object_id"]
        logger.info(f"  Archiva (Excel): {len(df):,} righe")
        return df.values.tolist()
    except Exception as e:
        logger.warning(f"Archiva Excel pandas failed ({e}), openpyxl fallback")
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb["Estrazione Archiva"]
        cols = [7, 94, 97, 115]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0: continue
            rows.append([row[c] if c < len(row) else None for c in cols])
        wb.close()
        return rows


def _load_archiva_csv(csv_path: str) -> list:
    """
    CSV cols: 0=r_object_id, 9=file_name, 96=PIVA, 99=RAGIONE_SOCIALE
    → list of [nome_file, piva, ragione_sociale, r_object_id]
    """
    logger.info(f"Loading Archiva from CSV: {csv_path} …")
    try:
        import pandas as pd
        df = pd.read_csv(csv_path, usecols=[0, 9, 96, 99],
                         header=0, dtype=object, sep=None, engine="python")
        # columns come in positional order: r_object_id, file_name, PIVA, RAGIONE_SOCIALE
        c = df.columns.tolist()
        df = df[[c[1], c[2], c[3], c[0]]]   # → nome_file, PIVA, RAGIONE_SOCIALE, r_object_id
        df.columns = ["nome_file", "piva", "ragione_sociale", "r_object_id"]
        logger.info(f"  Archiva (CSV): {len(df):,} righe")
        return df.values.tolist()
    except Exception as e:
        logger.error(f"CSV Archiva load failed: {e}")
        return []


# ── Record builder ────────────────────────────────────────────────────────────

def _build_records(rows_archiva: list, rows_postel: list) -> list:
    archiva_map: dict[str, dict] = {}

    for row in rows_archiva:
        nf, pv, rs, rid = row
        rid = _clean_str(rid)
        if not rid: continue
        nf = _clean_str(nf); pv = _clean_str(pv); rs = _clean_str(rs)
        dt = parse_date_from_filename(nf)
        archiva_map[rid] = {
            "r_object_id": rid, "nome_file": nf, "piva": pv, "ragione_sociale": rs,
            "data_doc":  dt.isoformat() if dt else "",
            "anno_mese": f"{dt.year:04d}/{dt.month:02d}" if dt else "",
            "in_archiva": "SI", "in_postel": "NO",
        }

    for row in rows_postel:
        rid, nf, rs, pv = row
        rid = _clean_str(rid)
        if not rid: continue
        nf = _clean_str(nf); pv = _clean_str(pv); rs = _clean_str(rs)
        dt = parse_date_from_filename(nf)

        if rid in archiva_map:
            archiva_map[rid]["in_postel"] = "SI"
            if not archiva_map[rid]["piva"]:          archiva_map[rid]["piva"] = pv
            if not archiva_map[rid]["ragione_sociale"]: archiva_map[rid]["ragione_sociale"] = rs
        else:
            archiva_map[rid] = {
                "r_object_id": rid, "nome_file": nf, "piva": pv, "ragione_sociale": rs,
                "data_doc":  dt.isoformat() if dt else "",
                "anno_mese": f"{dt.year:04d}/{dt.month:02d}" if dt else "",
                "in_archiva": "NO", "in_postel": "SI",
            }

    records = list(archiva_map.values())
    n_a = sum(1 for r in records if r["in_archiva"] == "SI")
    n_p = sum(1 for r in records if r["in_postel"] == "SI")
    n_e = sum(1 for r in records if r["in_archiva"] == "SI" and r["in_postel"] == "SI")
    logger.info(f"Records: total={len(records):,} archiva={n_a:,} postel={n_p:,} entrambi={n_e:,}")
    return records


# ── Search & stats ────────────────────────────────────────────────────────────

def search_records(records, ragione_sociale=None, piva=None, nome_file=None,
                   file_match="partial", date_from=None, date_to=None,
                   anno=None, mese=None) -> list:
    rs_q   = ragione_sociale.strip().lower() if ragione_sociale else None
    pv_q   = piva.strip().lower()            if piva            else None
    nf_q   = nome_file.strip().lower()       if nome_file       else None
    dt_f   = date_from.strip()               if date_from       else None
    dt_t   = date_to.strip()                 if date_to         else None
    anno_q = anno.strip().zfill(4)           if anno            else None
    mese_q = mese.strip().zfill(2)           if mese            else None

    out = []
    for r in records:
        if rs_q and rs_q not in r["ragione_sociale"].lower(): continue
        if pv_q and pv_q not in r["piva"].lower():           continue
        if nf_q:
            fn = r["nome_file"].lower()
            if file_match == "exact":
                if nf_q != fn: continue
            else:
                if nf_q not in fn: continue
        dd = r["data_doc"]
        if (dt_f or dt_t):
            if not dd: continue
            if dt_f and dd < dt_f: continue
            if dt_t and dd > dt_t: continue
        if anno_q and (not dd or not dd.startswith(anno_q)):  continue
        if mese_q and (not dd or dd[5:7] != mese_q):         continue
        out.append(r)
    return out


def compute_stats(records: list) -> dict:
    total    = len(records)
    archiva  = sum(1 for r in records if r["in_archiva"] == "SI")
    postel   = sum(1 for r in records if r["in_postel"]  == "SI")
    entrambi = sum(1 for r in records if r["in_archiva"] == "SI" and r["in_postel"] == "SI")
    return {
        "total":        total,
        "in_archiva":   archiva,
        "in_postel":    postel,
        "entrambi":     entrambi,
        "solo_archiva": archiva  - entrambi,
        "solo_postel":  postel   - entrambi,
        "quadratura":   round(entrambi / postel * 100, 1) if postel else 0.0,
    }
